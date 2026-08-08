import io
import os
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment

class StarkReportGenerator:
    """Motor de Generación de Reportes Técnicos Nivel Stark (Arquitectura Dual)."""

    def __init__(self, templates_dir="static/templates"):
        self.templates_dir = templates_dir
        # Buscamos primero en raíz, luego en templates_dir
        self.tec02_path = self._find_template("TEC-02_templates.xlsx")
        self.tec02a_path = self._find_template("tec02-A_template.xlsx")

    def _find_template(self, filename):
        """Busca en raíz y luego en static/templates."""
        if os.path.exists(filename):
            return filename
        return os.path.join(self.templates_dir, filename)

    def _get_master_cell_coord(self, sheet, cell_ref):
        for merged_range in sheet.merged_cells.ranges:
            if cell_ref in merged_range:
                return merged_range.coord.split(':')[0]
        return cell_ref

    def _safe_write(self, sheet, cell_ref, value, auto_height=False, chars_per_line=130):
        try:
            master_ref = self._get_master_cell_coord(sheet, cell_ref)
            sheet[master_ref] = value
            if auto_height and value:
                from copy import copy
                old_align = sheet[master_ref].alignment
                new_align = copy(old_align) if old_align else Alignment()
                new_align.wrap_text = True
                new_align.vertical = 'center'
                sheet[master_ref].alignment = new_align
                
                lines = str(value).split('\n')
                total_lines = 0
                for line in lines:
                    wraps = len(line) // chars_per_line
                    total_lines += (1 + wraps)
                row_idx = openpyxl.utils.cell.coordinate_from_string(master_ref)[1]
                # Excel default row height is ~15. Adding padding.
                sheet.row_dimensions[row_idx].height = max(15, total_lines * 15 + 10)
        except Exception:
            pass

    def _write_multiline(self, sheet, start_row, col_letter, text, max_rows=8, auto_height=True):
        lines = [line.rstrip('\r') for line in str(text).split('\n') if line.strip()] if text else []
        added_rows = 0
        current_row = start_row
        
        from copy import copy
        total_rows = max(len(lines), max_rows)
        
        for i in range(total_rows):
            if i >= max_rows:
                insert_idx = current_row
                sheet.insert_rows(insert_idx, 1)
                added_rows += 1
                
                # Copy style from row above
                for col in range(2, 28):
                    source_cell = sheet.cell(row=insert_idx - 1, column=col)
                    target_cell = sheet.cell(row=insert_idx, column=col)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
            
            # Openpyxl insert_rows corrupts merged cells and creates duplicates. 
            # We MUST cleanly rebuild the merged_cells.ranges list without any overlapping ranges for this row.
            new_ranges = []
            for mr in sheet.merged_cells.ranges:
                row_overlap = (mr.min_row == current_row and mr.max_row == current_row)
                col_overlap = not (mr.max_col < 2 or mr.min_col > 27)
                if not (row_overlap and col_overlap):
                    new_ranges.append(mr)
            sheet.merged_cells.ranges = new_ranges
                
            # Now explicitly merge B to AA
            try:
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=27)
            except Exception:
                pass
            
            if i < len(lines):
                line = lines[i]
                cell_ref = f"{col_letter}{current_row}"
                self._safe_write(sheet, cell_ref, line, auto_height=auto_height)
                
            current_row += 1
            
        return added_rows

    def _safe_write_rc(self, sheet, row, col, value):
        coord = f"{get_column_letter(col)}{row}"
        self._safe_write(sheet, coord, value)

    def generate_tec02_summary(self, candidates):
        """Genera un RESUMEN TABULAR (TEC-02) con todos los candidatos como filas."""
        wb = openpyxl.load_workbook(self.tec02_path)
        sheet = wb.active
        
        # HEADER STARK (Razón Social / Representante / Fecha)
        self._safe_write(sheet, "H9", "SERCOING LTDA")
        self._safe_write(sheet, "H11", "GUIDO CORTES ORDENES")
        
        # Fecha en formato Stark (D-M-YYYY)
        from datetime import datetime
        fecha_stark = datetime.now().strftime("%d-%m-%Y")
        self._safe_write(sheet, "W11", fecha_stark)

        start_row = 17
        for i, cand in enumerate(candidates):
            current_row = start_row + i
            
            def get_f(k, d=""):
                val = cand.get(k)
                if val is None: return d
                s = str(val).strip()
                if s.upper() in ["NONE", "NULL", "UNDEFINED", "N/A", ""]: return d
                return s

            c_nombre = get_f("nombre_completo", f"CANDIDATO_{i+1}")
            c_cargo = get_f("cargo_a_desempenar") or get_f("cargo") or "PERSONAL CLAVE"
            c_prof = get_f("profesion") or c_cargo or "TÉCNICO / PROFESIONAL ESPECIALIZADO"

            # Col B (Nº)
            self._safe_write_rc(sheet, current_row, 2, i + 1)
            # Col C (NOMBRE COMPLETO)
            self._safe_write_rc(sheet, current_row, 3, c_nombre.upper())
            # Col J (CARGO A DESEMPEÑAR)
            self._safe_write_rc(sheet, current_row, 10, c_cargo.upper())
            # Col O (TÍTULO PROFESIONAL)
            self._safe_write_rc(sheet, current_row, 15, c_prof.upper())
            # Col T (AÑOS EXP TOTAL - A)
            self._safe_write_rc(sheet, current_row, 20, cand.get("experiencia_total") or 0)
            # Col V (EXP EN LA EMPRESA - B)
            self._safe_write_rc(sheet, current_row, 22, cand.get("experiencia_en_empresa_actual") or 0)
            # Col X (EXP EN EL CARGO - C)
            self._safe_write_rc(sheet, current_row, 24, cand.get("exp_cargo_actual") or 0)
             # Col Z (EXP PROYECTOS SIMILARES - D)
            self._safe_write_rc(sheet, current_row, 26, cand.get("exp_proy_similares") or 0)

        target = io.BytesIO()
        wb.save(target)
        target.seek(0)
        return target

    def generate_tec02a_workbook(self, candidates):
        """Genera un archivo con MÚLTIPLES HOJAS (TEC-02A), una por cada perfil de candidato."""
        import copy
        from openpyxl.drawing.image import Image

        wb = openpyxl.load_workbook(self.tec02a_path)
        template_sheet = wb.active
        original_title = template_sheet.title
        
        # Extraer imágenes (logo CODELCO, etc.) de la hoja plantilla
        template_images = []
        for img in template_sheet._images:
            try:
                img_bytes = img._data()
                template_images.append((img_bytes, copy.deepcopy(img.anchor)))
            except Exception as ie:
                print(f"Advertencia al extraer imagen de la plantilla: {ie}")

        for cand in candidates:
            # Clonar la pestaña de perfil
            new_sheet = wb.copy_worksheet(template_sheet)
            
            # Copiar las imágenes (logo CODELCO, etc.) a la nueva hoja clonada
            for img_bytes, img_anchor in template_images:
                try:
                    new_img = Image(io.BytesIO(img_bytes))
                    new_img.anchor = copy.deepcopy(img_anchor)
                    new_sheet.add_image(new_img)
                except Exception as ie:
                    print(f"Advertencia al agregar imagen a la nueva hoja: {ie}")
            
            # Nombrar la pestaña (máx 31 caracteres)
            nombre = str(cand.get("nombre_completo", "Candidato"))[:25].strip()
            new_sheet.title = f"{nombre}_{str(cand.get('id', ''))[:4]}"

            # Sanitización helper para asegurar cero 'NONE', cero 'NULL' y cero vacíos
            def get_field(key, default=""):
                val = cand.get(key)
                if val is None: return default
                s = str(val).strip()
                if s.upper() in ["NONE", "NULL", "UNDEFINED", "N/A", ""]: return default
                return s

            nombre_cand = get_field("nombre_completo", "CANDIDATO SELECCIONADO")
            cargo_dest = get_field("cargo_a_desempenar") or get_field("cargo") or "PERSONAL CLAVE"
            profesion_cand = get_field("profesion") or cargo_dest or "TÉCNICO / PROFESIONAL ESPECIALIZADO"

            # Inyectar datos en la ficha individual (Anexo TEC-02A Stark)
            # ---------------------------------------------------------
            # H10: Razón Social
            self._safe_write(new_sheet, "H10", "SERCOING LTDA")
            # H12: Representante Legal
            self._safe_write(new_sheet, "H12", "GUIDO CORTES ORDENES")
            # W12: Fecha
            from datetime import datetime
            fecha_stark = datetime.now().strftime("%d-%m-%Y")
            self._safe_write(new_sheet, "W12", fecha_stark)

            # H17: Nombre
            self._safe_write(new_sheet, "H17", nombre_cand.upper())
            # H19: Título Profesional
            self._safe_write(new_sheet, "H19", profesion_cand.upper())
            # H21: Cargo Destino
            self._safe_write(new_sheet, "H21", cargo_dest.upper())

            # 1. BLOQUE EXPERIENCIA GENERAL (Header Row 23, Target B24 to B31 -> max 8 rows)
            exp_gen = get_field("experiencia_general")
            if not exp_gen:
                latest_emp = get_field("ultima_exp_laboral_empresa") or "SECTOR INDUSTRIAL"
                period = get_field("periodo") or "2018-PRESENTE"
                exp_gen = f"{period} {cargo_dest.upper()} - {latest_emp.upper()} - FAENA/OPERACIONES"
            added1 = self._write_multiline(new_sheet, 24, "B", exp_gen, max_rows=8, auto_height=True)
            offset1 = added1

            # 2. BLOQUE EXPERIENCIA ESPECÍFICA (Header Row 31, Target B33 to B39 -> max 7 rows)
            exp_esp = get_field("experiencia_especifica") or exp_gen
            added2 = self._write_multiline(new_sheet, 33 + offset1, "B", exp_esp, max_rows=7, auto_height=True)
            offset2 = offset1 + added2
            
            # 3. BLOQUE OTRAS EXPERIENCIAS (Header Row 40, Target B41 to B47 -> max 7 rows)
            exp_otras = get_field("otras_experiencias") or get_field("software_que_domina") or "CURSOS DE CAPACITACIÓN Y COMPETENCIAS TÉCNICAS INDUSTRIALES"
            added3 = self._write_multiline(new_sheet, 41 + offset2, "B", exp_otras, max_rows=7, auto_height=True)
            offset3 = offset2 + added3
            
            # 4. BLOQUE ANTECEDENTES ACADÉMICOS (Header Row 48, Target B49 to B55 -> max 7 rows)
            aca = get_field("antecedentes_academicos") or f"{profesion_cand.upper()} - INSTITUCIÓN EDUCACIÓN TÉCNICA / SUPERIOR - TITULADO"
            added4 = self._write_multiline(new_sheet, 49 + offset3, "B", aca, max_rows=7, auto_height=True)
            offset4 = offset3 + added4
            
            # 5. Fix ALL merges that openpyxl insert_rows corrupted (Headers + Bottom Block)
            merges_to_enforce = [
                # Headers
                (23, 2, 23, 27), # Exp General
                (32 + offset1, 2, 32 + offset1, 27), # Exp Especifica
                (40 + offset2, 2, 40 + offset2, 27), # Otras Exp
                (48 + offset3, 2, 48 + offset3, 27), # Antecedentes Acad
                
                # Notes
                (65 + offset4, 3, 66 + offset4, 26)   # C:Z Notes
            ]
            
            # Dinámicamente detectar la fila del bloque de firma en la plantilla original
            template_sig_row = 60
            for r in range(40, 100):
                if template_sheet.cell(row=r, column=12).value and "Nombre, Fecha" in str(template_sheet.cell(row=r, column=12).value):
                    template_sig_row = r
                    break
            
            template_name_row = template_sig_row - 1
            target_name_row = template_name_row + offset4
            
            # Dinámicamente detectar las celdas combinadas de la fila del nombre en la plantilla original
            for mr in template_sheet.merged_cells.ranges:
                if mr.min_row == template_name_row and mr.max_row == template_name_row:
                    merges_to_enforce.append((target_name_row, mr.min_col, target_name_row, mr.max_col))
            
            for r_min, c_min, r_max, c_max in merges_to_enforce:
                # Completely rebuild ranges list to aggressively purge ANY overlaps and duplicates
                new_ranges = []
                for mr in new_sheet.merged_cells.ranges:
                    row_overlap = not (mr.max_row < r_min or mr.min_row > r_max)
                    col_overlap = not (mr.max_col < c_min or mr.min_col > c_max)
                    if not (row_overlap and col_overlap):
                        new_ranges.append(mr)
                new_sheet.merged_cells.ranges = new_ranges
                
                # Re-apply correct merge
                coord = f"{get_column_letter(c_min)}{r_min}:{get_column_letter(c_max)}{r_max}"
                try:
                    new_sheet.merge_cells(coord)
                except Exception:
                    pass
                
                # Re-center headers and signature labels
                if r_min in [23, 32 + offset1, 40 + offset2, 48 + offset3]:
                    master = new_sheet.cell(row=r_min, column=c_min)
                    master.alignment = Alignment(horizontal='center', vertical='center')
            
            # 6. Escribir Nombre y Fecha en la posición final correcta (dinámica)
            from datetime import datetime
            fecha_stark = datetime.now().strftime("%d-%m-%Y")
            
            # CRÍTICO: Asegurar que las filas objetivo NO estén ocultas.
            # openpyxl insert_rows no desplaza los row_dimensions, por lo que si una fila 
            # cae matemáticamente en una fila oculta de la plantilla (ej: 69), se volverá invisible.
            if target_name_row in new_sheet.row_dimensions:
                new_sheet.row_dimensions[target_name_row].hidden = False
                new_sheet.row_dimensions[target_name_row].height = 15
            if target_name_row + 1 in new_sheet.row_dimensions:
                new_sheet.row_dimensions[target_name_row + 1].hidden = False
                new_sheet.row_dimensions[target_name_row + 1].height = 15
            
            self._safe_write(new_sheet, f"E{target_name_row}", str(cand.get("nombre_completo", "")).upper())
            self._safe_write(new_sheet, f"Q{target_name_row}", fecha_stark)
            
            # 7. Redibujar la línea de la firma unificada (openpyxl borra las formas insertadas)
            from openpyxl.styles import Border, Side
            line_border = Border(bottom=Side(style='thin', color='000000'))
            for c_idx in range(5, 25): # Columnas E(5) hasta X(24)
                new_sheet.cell(row=target_name_row, column=c_idx).border = line_border
                
            # 8. Limitar estrictamente el área de impresión para evitar páginas fantasma
            # El documento termina en las "NOTAS" (fila 66 aprox). Damos un margen hasta la 75.
            final_row = 105 + offset4
            new_sheet.print_area = f"A1:AA{final_row}"
            
        # Borrar la hoja original de plantilla
        if len(wb.sheetnames) > 1:
            wb.remove(wb[original_title])
            
        target = io.BytesIO()
        wb.save(target)
        target.seek(0)
        return target
