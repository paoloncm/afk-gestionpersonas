import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()
ws = wb.active

ws.merge_cells('B2:AA2')
ws['B2'] = 'Merged B2:AA2'
ws['B2'].alignment = Alignment(wrap_text=True)

ws.insert_rows(3, 1)
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=27)
ws['B3'] = 'Merged B3:AA3 inserted'
ws['B3'].alignment = Alignment(wrap_text=True)

wb.save("tmp/test_merge_insert.xlsx")
print("Saved tmp/test_merge_insert.xlsx")
