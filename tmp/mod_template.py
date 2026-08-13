import openpyxl
from openpyxl.styles import Border, Side, Alignment

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

# Clear old labels
ws['D60'] = None
ws['N60'] = None
ws['W60'] = None

# Clear old borders on row 59
for col in range(2, 28): # B to AA
    cell = ws.cell(row=59, column=col)
    cell.border = Border(bottom=None)

# Add single continuous line from D to Z
thin = Side(border_style="thin", color="000000")
for col in range(4, 27): # D (4) to Z (26)
    cell = ws.cell(row=59, column=col)
    cell.border = Border(bottom=thin)

# Add the new label
ws['D60'] = "Nombre, Fecha y Firma del Profesional"
# Center it across selection D to Z
ws.merge_cells("D60:Z60")
ws['D60'].alignment = Alignment(horizontal='center', vertical='center')

# Also, since we changed the merges, let's clean up B59:H59 and K59:Q59?
# If we keep them, the auto-fill will use B59 for Name and K59 for Date. That's probably fine.
# But if there's a single line from D to Z, maybe the Name should be placed differently?
# Let's leave B59 and K59 merges for now, we'll see how it looks.

wb.save("static/templates/tec02-A_template.xlsx")
print("Template updated")
