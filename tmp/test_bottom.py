import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.merge_cells('B59:H59')
ws['B59'] = "HUGO ARMANDO MANCINELLI LÓPEZ"

ws.insert_rows(32, 5)

# Simulate what report_gen.py does
t_r_min = 64
t_r_max = 64
c_min = 2
c_max = 8

ranges_to_remove = []
for mr in ws.merged_cells.ranges:
    if (mr.min_row >= t_r_min and mr.max_row <= t_r_max) or \
       (mr.min_row <= t_r_max and mr.max_row >= t_r_min):
        ranges_to_remove.append(mr)

for mr in ranges_to_remove:
    ws.merged_cells.ranges.remove(mr)
    
ws.merge_cells(start_row=t_r_min, start_column=c_min, end_row=t_r_max, end_column=c_max)

wb.save("tmp/test_bottom_merge.xlsx")
print("Saved tmp/test_bottom_merge.xlsx")
