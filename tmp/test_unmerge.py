import openpyxl
from openpyxl.styles import Border, Side, Alignment

wb = openpyxl.load_workbook('static/templates/tec02-A_template.xlsx')
sheet = wb.active

def split_merged_block_to_rows(sheet, cell_coord, text_lines):
    # Find the merged cell
    target_range = None
    for merged_range in sheet.merged_cells.ranges:
        if cell_coord in merged_range:
            target_range = merged_range
            break
            
    if target_range:
        # Unmerge
        sheet.unmerge_cells(str(target_range))
        
        min_col, min_row, max_col, max_row = target_range.bounds
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # We process line by line
        for i in range(max_row - min_row + 1):
            current_row = min_row + i
            # Merge horizontally
            row_range = f"{openpyxl.utils.cell.get_column_letter(min_col)}{current_row}:{openpyxl.utils.cell.get_column_letter(max_col)}{current_row}"
            sheet.merge_cells(row_range)
            
            # Get the cell
            cell = sheet.cell(row=current_row, column=min_col)
            
            # Apply border to all cells in the new horizontal merge so it looks like a box
            for col in range(min_col, max_col + 1):
                sheet.cell(row=current_row, column=col).border = thin_border
            
            # Write text if available
            if i < len(text_lines):
                cell.value = text_lines[i]
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.value = ""

split_merged_block_to_rows(sheet, "B24", ["Exp 1: bla bla", "Exp 2: bla bla", "Exp 3: bla bla"])

wb.save("tmp/test_unmerge.xlsx")
print("Done")
