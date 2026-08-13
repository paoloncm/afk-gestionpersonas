import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

for row in range(50, 75):
    val = ws.cell(row=row, column=12).value
    if val and "Nombre" in str(val):
        print(f"Found signature label at row {row}")
        
    name_val = ws.cell(row=row, column=5).value
    if name_val:
        print(f"Row {row} Col 5: {name_val}")
