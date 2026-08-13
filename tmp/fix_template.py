import openpyxl
from copy import copy

file_path = "static/templates/tec02-A_template.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

ranges_to_fix = ["B24:AA31", "B33:AA39", "B41:AA47", "B49:AA55"]

for r in ranges_to_fix:
    try:
        # Save style from the master cell before unmerging
        start_cell, end_cell = r.split(':')
        master = sheet[start_cell]
        align = copy(master.alignment) if master.alignment else None
        font = copy(master.font) if master.font else None
        border = copy(master.border) if master.border else None
        fill = copy(master.fill) if master.fill else None
        
        sheet.unmerge_cells(r)
        
        start_col = "".join(filter(str.isalpha, start_cell))
        start_row = int("".join(filter(str.isdigit, start_cell)))
        end_col = "".join(filter(str.isalpha, end_cell))
        end_row = int("".join(filter(str.isdigit, end_cell)))
        
        for row in range(start_row, end_row + 1):
            row_range = f"{start_col}{row}:{end_col}{row}"
            sheet.merge_cells(row_range)
            
            # Apply styles to the new row master cell
            cell = sheet[f"{start_col}{row}"]
            if align: cell.alignment = align
            if font: cell.font = font
            if border: cell.border = border
            if fill: cell.fill = fill
            
    except Exception as e:
        print(f"Error processing range {r}: {e}")

wb.save(file_path)
print("Template fixed successfully.")
