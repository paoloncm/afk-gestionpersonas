import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

for col in range(2, 27):
    cell59 = ws.cell(row=59, column=col)
    cell60 = ws.cell(row=60, column=col)
    
    b59 = cell59.border
    b60 = cell60.border
    
    if b59.bottom.style or b59.top.style or b60.bottom.style or b60.top.style:
        print(f"Col {col}: 59(top={b59.top.style}, bottom={b59.bottom.style}), 60(top={b60.top.style}, bottom={b60.bottom.style})")
