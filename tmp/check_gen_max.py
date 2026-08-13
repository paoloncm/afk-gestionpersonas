import os
from glob import glob
import openpyxl

# find the latest generated report in tmp or wherever it's generated
# Wait, let's just generate a report right now to inspect it!
import json
import sys
sys.path.append('c:\\Users\\Paolo\\diseño-afk')
from report_gen import StarkReportGenerator

cand_data = {
    "nombre_completo": "CARLOS ANTONIO HONORES",
    "profesion": "INGENIERO",
    "cargo_a_desempenar": "TEST",
    "experiencia_general": "Exp 1\nExp 2",
    "experiencia_especifica": "Exp 3\nExp 4",
    "otras_experiencias": "Exp 5",
    "antecedentes_academicos": "Edu 1"
}

gen = StarkReportGenerator()
wb_bytes = gen.generate_tec02a_workbook([cand_data])
wb = openpyxl.load_workbook(wb_bytes)
ws = wb.worksheets[0]
print(f"Generated Max Row: {ws.max_row}")
print(f"Generated Print Area: {ws.print_area}")

# Also check for empty rows that might be formatted?
count = 0
for row in ws.iter_rows(min_row=100):
    for cell in row:
        if cell.value is not None or cell.has_style:
            count += 1
            break
print(f"Rows > 100 with content or style: {count}")
