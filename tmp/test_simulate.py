import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
sheet = wb.active

# Simulate inserting 20 rows at row 32
sheet.insert_rows(32, 20)

offset = 20

bottom_merges = [
    (59, 2, 59, 8),    # B:H
    (59, 11, 59, 17),  # K:Q
    (65, 3, 66, 26)    # C:Z
]

for r_min, c_min, r_max, c_max in bottom_merges:
    t_r_min = r_min + offset
    t_r_max = r_max + offset
    
    ranges_to_remove = []
    for mr in sheet.merged_cells.ranges:
        if (mr.min_row >= t_r_min and mr.max_row <= t_r_max) or \
           (mr.min_row <= t_r_max and mr.max_row >= t_r_min):
            ranges_to_remove.append(mr)
            
    for mr in ranges_to_remove:
        sheet.merged_cells.ranges.remove(mr)
        
    coord = f"{openpyxl.utils.get_column_letter(c_min)}{t_r_min}:{openpyxl.utils.get_column_letter(c_max)}{t_r_max}"
    print(f"Merging {coord}")
    sheet.merge_cells(coord)

wb.save("tmp/test_simulate_failure.xlsx")
print("Saved")
