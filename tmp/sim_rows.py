import sys
sys.path.append('c:\\Users\\Paolo\\diseño-afk')
from report_gen import StarkReportGenerator
import openpyxl

candidates = [
    {
        "nombre_completo": "Yerson Erwin Arroyo Diaz",
        "experiencia_general": "Exp 1\nExp 2\nExp 3\nExp 4\nExp 5\nExp 6\nExp 7\nExp 8\nExp 9\nExp 10\nExp 11\nExp 12",
        "experiencia_especifica": "Esp 1",
        "otras_experiencias": "Otras 1",
        "antecedentes_academicos": "Acad 1"
    },
    {
        "nombre_completo": "Alberto Enrique Escudero Zamora",
        "experiencia_general": "Exp 1",
        "experiencia_especifica": "Esp 1",
        "otras_experiencias": "Otras 1",
        "antecedentes_academicos": "Técnico Mecánico Automotriz Diesel - Fuerza Aérea De Chile - Titulado\nTécnico En Enfermería - En Curso"
    }
]

gen = StarkReportGenerator()
wb_bytes = gen.generate_tec02a_workbook(candidates)
wb = openpyxl.load_workbook(wb_bytes)

for sheet in wb.worksheets:
    print(f"--- Sheet: {sheet.title} ---")
    # Find signature block
    for row in range(40, 100):
        val_l = sheet.cell(row=row, column=12).value
        val_e = sheet.cell(row=row, column=5).value
        if val_l and "Nombre, Fecha" in str(val_l):
            print(f"  Signature label 'Nombre, Fecha' is at row {row}")
        if val_e and ("YERSON" in str(val_e) or "ALBERTO" in str(val_e)):
            print(f"  Candidate Name is at row {row}")
