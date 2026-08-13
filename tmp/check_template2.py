import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

for row in range(56, 62):
    for col in range(2, 27):
        val = ws.cell(row=row, column=col).value
        if val:
            print(f"Row {row}, Col {col} ({openpyxl.utils.get_column_letter(col)}{row}): {val}")
