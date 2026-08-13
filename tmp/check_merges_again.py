import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

for mr in ws.merged_cells.ranges:
    if mr.min_row == 59 and mr.max_row == 59:
        print(f"Row 59 merge: {mr.coord}")
    if mr.min_row == 60 and mr.max_row == 60:
        print(f"Row 60 merge: {mr.coord}")
