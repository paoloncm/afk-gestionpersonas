import openpyxl

wb = openpyxl.load_workbook("static/templates/tec02-A_template.xlsx")
ws = wb.active

for col in range(2, 27):
    val = ws.cell(row=60, column=col).value
    if val:
        print(f"Col {col} (Row 60): {val}")
        
for col in range(2, 27):
    val = ws.cell(row=59, column=col).value
    if val:
        print(f"Col {col} (Row 59): {val}")
